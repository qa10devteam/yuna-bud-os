"""Faza 45 — Excel Import/Export: upload XLSX do tenderów, eksport pipeline."""
from __future__ import annotations


import io
import json
import uuid
from datetime import datetime
from typing import Any

import sqlalchemy as sa
from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse

from terra_db.session import get_engine
from ..auth.deps import AuthUser

router = APIRouter(prefix="/api/v1/excel", tags=["excel"])


def _process_xlsx_tenders(content: bytes, org_id: str) -> tuple[int, list[str]]:
    """Parse XLSX and import tenders. Returns (rows_imported, errors)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        if ws is None:
            return 0, ["Brak arkusza w pliku XLSX"]

        engine = get_engine()
        imported = 0
        errors = []
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            try:
                row_data = dict(zip(headers, row))
                title = str(row_data.get("title") or row_data.get("tytuł") or row_data.get("nazwa") or "")
                buyer = str(row_data.get("buyer") or row_data.get("zamawiający") or "")
                value = row_data.get("value_pln") or row_data.get("wartość") or row_data.get("value")
                try:
                    value_pln = float(str(value).replace(",", ".").replace(" ", "")) if value else None
                except (ValueError, TypeError):
                    value_pln = None

                if not title:
                    errors.append(f"Wiersz {row_idx}: brak tytułu")
                    continue

                with engine.connect() as conn:
                    conn.execute(
                        sa.text("""
                            INSERT INTO tender (id, title, buyer, value_pln, source, status, created_at)
                            VALUES (:id, :title, :buyer, :value, 'excel', 'new', now())
                        """),
                        {
                            "id": str(uuid.uuid4()),
                            "title": title[:500],
                            "buyer": buyer[:300],
                            "value": value_pln,
                        },
                    )
                    conn.commit()
                imported += 1
            except Exception as e:
                errors.append(f"Wiersz {row_idx}: {e}")
        return imported, errors
    except Exception as e:
        return 0, [f"Błąd parsowania XLSX: {e}"]


def _export_tenders_xlsx(org_id: str | None) -> bytes:
    """Export tenders to XLSX bytes."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Przetargi"

        headers = ["ID", "Tytuł", "Zamawiający", "Status", "CPV", "Wartość PLN", "Termin", "Źródło", "Data dodania"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        engine = get_engine()
        with engine.connect() as conn:
            rows = conn.execute(
                sa.text("""
                    SELECT id, title, buyer, status, cpv, value_pln, deadline_at, source, created_at
                    FROM tender
                    ORDER BY created_at DESC
                    LIMIT 5000
                """),
            ).fetchall()

        for row_idx, r in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=str(r.id))
            ws.cell(row=row_idx, column=2, value=r.title)
            ws.cell(row=row_idx, column=3, value=r.buyer)
            ws.cell(row=row_idx, column=4, value=r.status)
            ws.cell(row=row_idx, column=5, value=",".join(r.cpv) if r.cpv else "")
            ws.cell(row=row_idx, column=6, value=float(r.value_pln) if r.value_pln else "")
            ws.cell(row=row_idx, column=7, value=r.deadline_at.strftime("%Y-%m-%d") if r.deadline_at else "")
            ws.cell(row=row_idx, column=8, value=r.source)
            ws.cell(row=row_idx, column=9, value=r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "")

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")


@router.post("/import/tenders")
async def import_tenders_xlsx(
    user: AuthUser,
    file: UploadFile = File(...),
) -> dict:
    """Import przetargów z pliku XLSX."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Wymagany plik XLSX")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:  # 10MB limit
        raise HTTPException(status_code=400, detail="Plik zbyt duży (maks 10MB)")

    engine = get_engine()
    import_id = str(uuid.uuid4())
    org_id = user.org_id or ""

    # Record import attempt
    with engine.connect() as conn:
        conn.execute(
            sa.text("""
                INSERT INTO excel_imports (id, org_id, user_id, filename, import_type, status)
                VALUES (:id, :org_id, :user_id, :filename, 'tender', 'processing')
            """),
            {
                "id": import_id,
                "org_id": org_id or None,
                "user_id": user.user_id,
                "filename": file.filename,
            },
        )
        conn.commit()

    imported, errors = _process_xlsx_tenders(content, org_id)

    with engine.connect() as conn:
        conn.execute(
            sa.text("""
                UPDATE excel_imports SET rows_imported=:rows, errors=cast(:errs as jsonb), status=:status
                WHERE id=:id
            """),
            {
                "id": import_id,
                "rows": imported,
                "errs": json.dumps(errors),
                "status": "done" if not errors else "done_with_errors",
            },
        )
        conn.commit()

    return {
        "import_id": import_id,
        "filename": file.filename,
        "rows_imported": imported,
        "errors": errors[:20],
        "status": "done",
    }


@router.get("/export/tenders")
def export_tenders_xlsx(user: AuthUser) -> StreamingResponse:
    """Eksport przetargów do pliku XLSX."""
    content = _export_tenders_xlsx(user.org_id)
    filename = f"przetargi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/imports")
def list_imports(user: AuthUser, limit: int = Query(50)) -> dict:
    """Historia importów XLSX."""
    engine = get_engine()
    with engine.connect() as conn:
        rows = conn.execute(
            sa.text("""
                SELECT id, filename, import_type, rows_imported, errors, status, created_at
                FROM excel_imports
                ORDER BY created_at DESC LIMIT :limit
            """),
            {"limit": limit},
        ).fetchall()
    return {
        "items": [
            {
                "id": str(r.id),
                "filename": r.filename,
                "import_type": r.import_type,
                "rows_imported": r.rows_imported,
                "status": r.status,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in rows
        ]
    }
