"""S47 — Import historii ofert z pliku XLS/XLSX.

Endpoint POST /api/v2/offers/import-history (multipart file).
Kolumny: nr_postepowania, status, kwota_oferty, data_zlozone.
Match po noticeNumber z tabeli tender. Upsert do offer_result.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Any

import sqlalchemy as sa
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import JSONResponse

from terra_db.session import get_engine
from ..auth.deps import AuthUser, get_current_user

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v2/offers", tags=["offer-history"])

_STATUS_MAP = {
    "wygrany": "won", "wygrana": "won", "won": "won", "win": "won", "1": "won",
    "przegrany": "lost", "przegrana": "lost", "lost": "lost", "lose": "lost", "0": "lost",
    "anulowany": "cancelled", "cancelled": "cancelled", "cancel": "cancelled",
    "wycofany": "withdrawn", "withdrawn": "withdrawn",
    "złożony": "submitted", "zlozone": "submitted", "submitted": "submitted",
}


def _parse_date(val: Any) -> datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if hasattr(val, "date"):
        return datetime(val.year, val.month, val.day)
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _parse_float(val: Any) -> float | None:
    if val is None:
        return None
    try:
        return float(str(val).replace(",", ".").replace(" ", "").replace("\xa0", ""))
    except (ValueError, TypeError):
        return None


@router.post("/import-history")
def import_offer_history(file: UploadFile = File(...), user: AuthUser = Depends(get_current_user)) -> dict:
    """Import historii wyników ofert z pliku XLSX."""
    if not user or not user.org_id:
        raise HTTPException(status_code=403, detail="Brak org_id")

    tenant_id = user.org_id

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Błąd parsowania pliku: {e}")

    # Read headers from first row
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    col_map = {}
    for i, h in enumerate(headers):
        if any(k in h for k in ["nr_post", "nr post", "numer post", "notice"]):
            col_map["notice_number"] = i
        elif any(k in h for k in ["status"]):
            col_map["status"] = i
        elif any(k in h for k in ["kwota", "value", "bid"]):
            col_map["bid_value"] = i
        elif any(k in h for k in ["data_zl", "data zl", "submitted", "złożenia"]):
            col_map["submitted_at"] = i
        elif any(k in h for k in ["data_dec", "decided", "wynik"]):
            col_map["decided_at"] = i
        elif any(k in h for k in ["konkur", "compet", "rywal"]):
            col_map["competitor_name"] = i
        elif any(k in h for k in ["uwagi", "notes", "notatki"]):
            col_map["notes"] = i

    engine = get_engine()
    imported = 0
    skipped = 0
    errors = []

    with engine.begin() as conn:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            try:
                notice_number = str(row[col_map["notice_number"]]).strip() if "notice_number" in col_map else None
                if not notice_number:
                    skipped += 1
                    continue

                # Find tender_id
                tender_row = conn.execute(
                    sa.text("SELECT id FROM tender WHERE notice_number = :nn LIMIT 1"),
                    {"nn": notice_number},
                ).fetchone()
                tender_id = str(tender_row[0]) if tender_row else None

                raw_status = str(row[col_map.get("status", -1)] or "").strip().lower() if "status" in col_map else ""
                status = _STATUS_MAP.get(raw_status, "submitted")
                bid_value = _parse_float(row[col_map["bid_value"]]) if "bid_value" in col_map else None
                submitted_at = _parse_date(row[col_map["submitted_at"]]) if "submitted_at" in col_map else None
                decided_at = _parse_date(row[col_map["decided_at"]]) if "decided_at" in col_map else None
                competitor_name = str(row[col_map["competitor_name"]] or "").strip() if "competitor_name" in col_map else None
                notes = str(row[col_map["notes"]] or "").strip() if "notes" in col_map else None

                # Upsert by notice_number + tenant_id
                conn.execute(
                    sa.text("""
                        INSERT INTO offer_result
                            (tenant_id, tender_id, status, bid_value_pln, submitted_at,
                             decided_at, competitor_name, notes, notice_number)
                        VALUES
                            (:tenant_id, :tender_id, :status, :bid_value, :submitted_at,
                             :decided_at, :competitor_name, :notes, :notice_number)
                        ON CONFLICT DO NOTHING
                    """),
                    {
                        "tenant_id": tenant_id,
                        "tender_id": tender_id,
                        "status": status,
                        "bid_value": bid_value,
                        "submitted_at": submitted_at,
                        "decided_at": decided_at,
                        "competitor_name": competitor_name,
                        "notes": notes,
                        "notice_number": notice_number,
                    },
                )
                imported += 1
            except Exception as e:
                errors.append(f"Wiersz {row_idx}: {e}")
                skipped += 1

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:20],
        "tenant_id": tenant_id,
    }
